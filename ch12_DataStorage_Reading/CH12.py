# CSV(Comma Separated Values)以逗點分隔值的資料格式，最廣泛運用
# csv.writer &  csv.reader
# test1 : 1D list write to CSV
import csv
# add newline='' 可以讓資料中換航自原被正確解析
with open('test1.csv','w',newline='') as csvfile:
    writer = csv.writer(csvfile)  #  # 建立 csv 檔寫入物件
    # writerow
    writer.writerow(['姓名', '身高', '體重'])  # 寫入欄位名稱
    writer.writerow(['Chiou', 170, 65])     # 寫入資料
    writer.writerow(['David', 183, 78])
# PS : 程式出了 with , csv file close
# test2 : 2D list write to csv ----------------------------------
import csv
# 建立csv二維串列資料
csvtable = [
        ['姓名', '身高', '體重'],
        ['Chiou', 170, 65],
        ['David', 183, 78],
]
with open('test2.csv', 'w', newline='') as csvfile:
  # 建立 csv 檔寫入物件
  writer = csv.writer(csvfile)
  # 寫入二維串列資料 : writerows
  writer.writerows(csvtable)
# test3 dictionary write to csv-----------------------------------
import csv
with open('test3.csv', 'w', newline='') as csvfile:
    # 定義欄位
    fieldnames = ['姓名', '身高', '體重']
    # 將 dictionary 寫入 csv 檔  csv.DictWriter
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    # 寫入欄位名稱
    writer.writeheader()
    # 寫入資料
    writer.writerow({'姓名': 'Chiou', '身高': 170, '體重': 65})
    writer.writerow({'姓名': 'David', '身高': 183, '體重': 78})
test4 CSV read --------------------------------------------------
import csv
# 開啟 csv 檔案
with open('test1.csv', newline='') as csvfile:
    # 讀取 csv 檔案內容
    rows = csv.reader(csvfile)
    # 以迴圈顯示每一列
    for row in rows:
        print(row)
# test5  read csv to dictionary -----------------------------------
import csv
# 開啟 csv 檔案
with open('test1.csv', newline='') as csvfile:
    # 讀取 csv 檔內容，將每一列轉成 dictionary:csv.DictReader
    rows = csv.DictReader(csvfile)
    # 以迴圈顯示每一列
    for row in rows:
        print(row['姓名'], row['身高'], row['體重'])


# 2.存取 Excel 文件格式 use openpyxl 模組
# test1 write Excel -------------------------
import openpyxl
# 建立一個工作簿
workbook=openpyxl.Workbook()
# 取得第 1 個工作表
sheet = workbook.worksheets[0]
# 以儲存格位置寫入資料
sheet['A1'] = 'Hello'
sheet['B1'] = 'World'
# 以串列寫入資料
listtitle=["姓名","電話"]
sheet.append(listtitle)
listdata=["David","0999-1234567"]
sheet.append(listdata)
# 儲存檔案
workbook.save('test.xlsx')
# test2 read Excel --------------------------
import openpyxl
#  讀取檔案
workbook = openpyxl.load_workbook('test.xlsx')
# 取得第 1 個工作表
sheet = workbook.worksheets[0]
# 取得指定儲存格
print(sheet['A1'], sheet['A1'].value)
# 取得總行、列數
print(sheet.max_row, sheet.max_column)
# 顯示 cell資料
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value,end="   ")
    print()
sheet['A3'] = 'Perry'
workbook.save('test.xlsx')

# 3. SQLite is python3 內建的崁入式小型資料庫 : 可以進行新增，修改，刪除，查詢
# 使用 sqlite3 模組
# import sqlite3
# conn = sqlite3.connect(資料庫檔案)  #回傳一個connection 物件
# conn.close()
# (1)cursor() : 建立一個cursor 物件，利用物件的execute 完成資料表的新噌修刪除&查詢
# (2)execute(SQL命令)
# (3)commit() 資料庫更新
# (4)close() 關閉資料庫連線
# test1 ------------------------------------------
import sqlite3
conn = sqlite3.connect('test.sqlite') # 建立資料庫連線
cursor = conn.cursor() # 建立 cursor 物件
# 建立一個資料表(SQL  語法: 不分大小寫)
sqlstr='''CREATE TABLE IF NOT EXISTS table01 \
("id"  INTEGER PRIMARY KEY NOT NULL,
 "name"  TEXT NOT NULL,
 "tel"  TEXT NOT NULL)
'''
cursor.execute(sqlstr)

# 新增一筆記錄
sqlstr='insert into table01 values(1, "David", "02-1234567")'
cursor.execute(sqlstr)
conn.commit() # 更新
conn.close()  # 關閉資料庫連線
# #----------------------------------------
# download DB Browser for SQLite 管理資料庫: https://sqlitebrowser.org/dl/
# install DB.Browser.for.SQLite-3.12.2-win64
# can check "test.sqlite" Databbase Name

# test2 新增資料表 ---------------------------------
# SQLite 資料型態 INTEGER,REAL(浮點數)，TEXT，BOLB(二進位資料)
import sqlite3
conn = sqlite3.connect('test.sqlite') # 建立資料庫連線
# 建立一個資料表
sqlstr='''CREATE TABLE "contact" \
("id"  INTEGER PRIMARY KEY NOT NULL,
 "name"  TEXT NOT NULL,
 "tel"  TEXT NOT NULL)
'''
conn.execute(sqlstr)
conn.commit() # 更新
conn.close()  # 關閉資料庫連線
# test3 新增資料 --------------------------------------
import sqlite3
conn = sqlite3.connect('test.sqlite') # 建立資料庫連線
# 定義資料串列
datas = [[1, 'David', '02-123456789'],
        [2, 'Lily', '02-987654321'],]
for data in datas:
    # 新增資料
    conn.execute("INSERT INTO contact (id, name, tel) VALUES \
                 ({}, '{}', '{}')".format(data[0], data[1], data[2]))
conn.commit() # 更新
conn.close()  # 關閉資料庫連線
# test3 更新資料 ------------------------------------------
import sqlite3
conn = sqlite3.connect('test.sqlite') # 建立資料庫連線
# 更新資料
conn.execute("UPDATE contact SET name='{}' WHERE id={}".format('Ken', 1))
conn.commit() # 更新
conn.close()  # 關閉資料庫連線

test4 刪除資料
import sqlite3
conn = sqlite3.connect('test.sqlite') # 建立資料庫連線
# 刪除資料
conn.execute("DELETE FROM contact WHERE id={}".format(1))
conn.commit() # 更新
conn.close()  # 關閉資料庫連線

# drop 資料表 : conn.execute("DROP TABLE contact")
# test5 資料查詢 fetchall() , fetchone() -----------------------------------
import sqlite3
conn = sqlite3.connect('test.sqlite') # 建立資料庫連線
cursor = conn.execute('select * from contact')
rows = cursor.fetchall()
# 顯示原始資料
print(rows)
# 逐筆顯示資料
for row in rows:
    print(row[0],row[1])
conn.close()  # 關閉資料庫連線
#-------------------------------------------------------------------
# test6
import sqlite3
conn = sqlite3.connect('test.sqlite') # 建立資料庫連線
cursor = conn.execute('select * from contact')
row = cursor.fetchone()
print(row[0], row[1])
conn.close()  # 關閉資料庫連

# 4. MySQL 資料庫操作 --------------------------------
# (1). MySql 安裝 : Download : https://dev.mysql.com/downloads/
# (2). Select : MySQL Community Server
# (3). 一定要安裝的是MySQL Server & MySQL Workbench
# (4).使用MySQL : Step 1 : 開啟 MySQL Shell : 開啟 & 關閉資料庫
# (5).  Step 2 : 開啟 MySQL Workbench : SQL 指令輸入環境 +管理工具
# (6). Install PyMySql 模組: pip install pymysql
# Create Scores Table --------------------------
import pymysql
conn = pymysql.connect(host="127.0.0.1", port=3306, user="root", password="James$981231", db="pythondb", charset="utf8")
with conn.cursor() as cursor:
    sql = """
    CREATE TABLE IF NOT EXISTS Scores(
      ID int NOT NULL AUTO_INCREMENT PRIMARY KEY,
      Name varchar(20),
      Chinese int(3),
      English int(3),
      Math int(3)
    );
    """
    cursor.execute(sql)
    conn.commit()
conn.close()
# 這種寫法比較好-----------------------------------------
import pymysql

# import charts
# 資料庫參數設定
db_settings = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "James$981231",
    "db": "pythondb",
    "charset": "utf8"
}
try:
    # 建立Connection物件
    conn = pymysql.connect(**db_settings)
    # 建立Cursor物件
    with conn.cursor() as cursor:
        # 新增資料SQL語法
        command = """
        CREATE TABLE IF NOT EXISTS Scores(
         ID int NOT NULL AUTO_INCREMENT PRIMARY KEY,
         Name varchar(20),
         Chinese int(3),
         English int(3),
         Math int(3)
        );
        """
        cursor.execute(command)
        conn.commit()
    conn.close()
except Exception as ex:
    print(ex)
# 新增 ，查詢 ，更新、刪除
# 新增資料    #  insert into tablename(欄位1，欄位2,...) values (值1, 值2,...)
# in score table 新增5筆datas -----------------------------------------------
import pymysql

db_settings = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "James$981231",
    "db": "pythondb",
    "charset": "utf8"
}
try:
    conn = pymysql.connect(**db_settings)  # 建立Connection物件
    with conn.cursor() as cursor:  # 建立Cursor物件
        command = """
        insert into scores(Name, Chinese, English, Math) values
        ('李大毛',95,92,80),
        ('林小明',82,83,61),
        ('黃小英',74,53,71),
        ('劉大樹',86,87,89),
        ('何美麗',89,73,95)        
        """
        cursor.execute(command)
        conn.commit()
    conn.close()
except Exception as ex:
    print(f'ex={ex}')
# 查詢資料  # select 欄位一，欄位二，.... from dtattable where 條件式 ---------------------
import pymysql

db_settings = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "James$981231",
    "db": "pythondb",
    "charset": "utf8"
}
try:
    conn = pymysql.connect(**db_settings)  # 建立Connection物件
    with conn.cursor() as cursor:  # 建立Cursor物件
        command = "Select * from scores"
        cursor.execute(command)
        datas = cursor.fetchall()  # datas is tuple
        print(datas)
        print("-" * 30)
        command = "Select * from scores"
        cursor.execute(command)
        data = cursor.fetchone()
        print(data)
        conn.commit()
    conn.close()
except Exception as ex:
    print(f'ex={ex}')

# 更新資料      # update 資料表 set 欄位1 = 值1, 欄位2 = 值2 ... 條件式-----------------------
import pymysql

db_settings = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "James$981231",
    "db": "pythondb",
    "charset": "utf8"
}
try:
    conn = pymysql.connect(**db_settings)  # 建立Connection物件
    with conn.cursor() as cursor:  # 建立Cursor物件
        command = "update scores set Chinese =98 where ID =4 "
        cursor.execute(command)
        conn.commit()
        command = "Select * from scores where ID=4"
        cursor.execute(command)
        data = cursor.fetchone()
        print(data)
        conn.commit()
    conn.close()
except Exception as ex:
    print(f'ex={ex}')

# 刪除資料   # delete form 資料表 where 條件式-------------------------------------
import pymysql

db_settings = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "James$981231",
    "db": "pythondb",
    "charset": "utf8"
}
try:
    conn = pymysql.connect(**db_settings)  # 建立Connection物件
    with conn.cursor() as cursor:  # 建立Cursor物件
        command = "delete from scores where ID =4 "
        cursor.execute(command)
        conn.commit()

        command = "select * from scores"
        cursor.execute(command)
        datas = cursor.fetchall()
        print(datas)
        conn.commit()
    conn.close()
except Exception as ex:
    print(f'ex={ex}')

# google 試算表  : https://console.developers.google.com
# step1 : 建立google 應用程式授權憑證: google developers Console
# 建立專案，啟動 Google Sheet API ，build 服務帳號，金鑰
# step2 :建立google 試算表，設權限給程式用
# step3 : 安裝gspread, oauth2client module
import gspread
from oauth2client.service_account import ServiceAccountCredentials as sac

auth_json = 'pythonconnectgsheet1-341814-13531c8bf732.json'  # 金鑰
gs_scopes = [
    'https://docs.google.com/spreadsheets/d/1pJ0MLUhGDK9Q753JakcBaLTrrfogg5lb8V738uP-og0/edit?usp=sharing']  # 要用google 試算表範圍
# 連線資料表
cr = sac.from_json_keyfile_name(auth_json, gs_scopes)
gs_client = gspread.authorize(cr)
# 開啟資料表
spreadsheet_key = 'pythonconnectgsheet1-341814'
sheet = gs_client.open_by_key(spreadsheet_key)
# 開啟工作簿
wks = sheet.sheet1
# 清除所有內容
wks.clear()
# 新增列
listtitle = ["姓名", "電話"]
wks.append_row(listtitle)  # 標題
listdata = ["chiou", "0937-1234567"]
wks.append_row(listdata)  # 資料

